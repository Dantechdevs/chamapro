from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Count
from decimal import Decimal
import datetime
import io
import csv
import json
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from .models import (
    User, Chama, Membership, Contribution, Penalty,
    Loan, LoanRepayment, MpesaTransaction,
    NotificationPreference, MemberActivity,
)
from .mpesa import mpesa


# ─── Auth ────────────────────────────────────────────────────────────────────

def home(request):
    return render(request, 'home.html')

def features(request):
    return render(request, 'features.html')

def pricing(request):
    return render(request, 'pricing.html')

def customers(request):
    return render(request, 'customers.html')

def forgot_password(request):
    return render(request, 'forgot_password.html')


def login(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        identity = request.POST.get('identity', '').strip()
        password = request.POST.get('password', '').strip()
        if not identity or not password:
            messages.error(request, 'Please enter your email/phone and password.')
            return render(request, 'login.html')
        user = None
        if '@' in identity:
            try:
                u = User.objects.get(email=identity)
                user = authenticate(request, username=u.username, password=password)
            except User.DoesNotExist:
                pass
        else:
            try:
                u = User.objects.get(phone=identity)
                user = authenticate(request, username=u.username, password=password)
            except User.DoesNotExist:
                user = authenticate(request, username=identity, password=password)
        if user:
            auth_login(request, user)
            return redirect(request.GET.get('next', 'dashboard'))
        messages.error(request, 'Invalid credentials. Please check and try again.')
    return render(request, 'login.html')


def signup(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        full_name = request.POST.get('full_name', '').strip()
        email     = request.POST.get('email', '').strip()
        phone     = request.POST.get('phone', '').strip()
        password  = request.POST.get('password', '').strip()
        errors = []
        if not full_name: errors.append('Full name is required.')
        if not email: errors.append('Email is required.')
        if not password or len(password) < 6: errors.append('Password must be at least 6 characters.')
        if User.objects.filter(email=email).exists(): errors.append('Email already registered.')
        if phone and User.objects.filter(phone=phone).exists(): errors.append('Phone already registered.')
        if errors:
            for e in errors: messages.error(request, e)
            return render(request, 'signup.html')
        name_parts = full_name.split(' ', 1)
        base = email.split('@')[0]
        username = base
        i = 1
        while User.objects.filter(username=username).exists():
            username = f'{base}{i}'; i += 1
        user = User.objects.create_user(
            username=username, email=email, password=password,
            first_name=name_parts[0], last_name=name_parts[1] if len(name_parts) > 1 else '',
            phone=phone or None,
        )
        auth_login(request, user)
        return redirect('dashboard')
    return render(request, 'signup.html')


def logout(request):
    auth_logout(request)
    return redirect('login')


# ─── Dashboard ────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def dashboard(request):
    memberships = request.user.memberships.filter(active=True).select_related('chama')
    chamas = [m.chama for m in memberships]
    active_chama_id = request.session.get('active_chama_id')
    active_chama = None
    active_membership = None
    if active_chama_id:
        active_chama = next((c for c in chamas if c.id == active_chama_id), None)
    if not active_chama and chamas:
        active_chama = chamas[0]
        request.session['active_chama_id'] = active_chama.id
    if active_chama:
        try:
            active_membership = Membership.objects.get(chama=active_chama, user=request.user)
        except Membership.DoesNotExist:
            pass
    context = {
        'user': request.user,
        'chamas': chamas,
        'active_chama': active_chama,
        'active_membership': active_membership,
    }
    return render(request, 'dashboard.html', context)


@login_required(login_url='login')
def switch_chama(request, chama_id):
    get_object_or_404(Membership, chama_id=chama_id, user=request.user, active=True)
    request.session['active_chama_id'] = chama_id
    return redirect('dashboard')


# ─── Profile ─────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def profile(request):
    user = request.user
    prefs, _ = NotificationPreference.objects.get_or_create(user=user)
    memberships = user.memberships.filter(active=True).select_related('chama').order_by('joined_at')

    chama_stats = []
    for m in memberships:
        paid = m.chama.contributions.filter(
            member=user, status='confirmed'
        ).aggregate(t=Sum('amount'))['t'] or Decimal('0')
        last = m.chama.contributions.filter(
            member=user, status='confirmed'
        ).order_by('-date').first()
        chama_stats.append({
            'membership': m,
            'chama': m.chama,
            'total_paid': paid,
            'last_contribution': last,
        })

    recent_activity = user.activities.select_related('chama').order_by('-created_at')[:10]

    today = datetime.date.today()
    monthly_trend = []
    for i in range(5, -1, -1):
        d = (today.replace(day=1) - datetime.timedelta(days=i * 30)).replace(day=1)
        month_total = user.contributions.filter(
            status='confirmed',
            date__month=d.month,
            date__year=d.year,
        ).aggregate(t=Sum('amount'))['t'] or Decimal('0')
        monthly_trend.append({'label': d.strftime('%b'), 'amount': float(month_total)})

    max_trend = max((m['amount'] for m in monthly_trend), default=1) or 1

    context = {
        'user': user,
        'prefs': prefs,
        'memberships': memberships,
        'chama_stats': chama_stats,
        'recent_activity': recent_activity,
        'monthly_trend': monthly_trend,
        'max_trend': max_trend,
        'credit_score': user.credit_score(),
        'credit_label': user.credit_score_label(),
        'credit_color': user.credit_score_color(),
        'credit_pct': int((user.credit_score() / 850) * 100),
        'payment_rate': user.payment_rate(),
        'total_contributed': user.total_contributed_all(),
        'active_loan_balance': user.active_loan_balance(),
        'outstanding_fines': user.outstanding_fines(),
        'chama_count': memberships.count(),
    }
    return render(request, 'profile.html', context)


@login_required(login_url='login')
def profile_edit(request):
    user = request.user
    prefs, _ = NotificationPreference.objects.get_or_create(user=user)

    if request.method == 'POST':
        user.first_name   = request.POST.get('first_name', '').strip()
        user.last_name    = request.POST.get('last_name', '').strip()
        user.email        = request.POST.get('email', '').strip()
        user.phone        = request.POST.get('phone', '').strip() or None
        user.national_id  = request.POST.get('national_id', '').strip() or None
        user.kra_pin      = request.POST.get('kra_pin', '').strip() or None
        user.occupation   = request.POST.get('occupation', '').strip() or None
        user.location     = request.POST.get('location', '').strip() or None
        user.bio          = request.POST.get('bio', '').strip() or None
        user.mpesa_number = request.POST.get('mpesa_number', '').strip() or None
        user.bank_name    = request.POST.get('bank_name', '').strip() or None

        bank_account = request.POST.get('bank_account', '').strip()
        if bank_account:
            user.bank_account = (
                'x' * (len(bank_account) - 4) + bank_account[-4:]
                if len(bank_account) > 4 else bank_account
            )

        if 'avatar' in request.FILES:
            user.avatar = request.FILES['avatar']

        user.save()

        prefs.mpesa_alerts     = request.POST.get('mpesa_alerts') == 'on'
        prefs.sms_reminders    = request.POST.get('sms_reminders') == 'on'
        prefs.email_reports    = request.POST.get('email_reports') == 'on'
        prefs.report_frequency = request.POST.get('report_frequency', 'weekly')
        prefs.two_fa_enabled   = request.POST.get('two_fa_enabled') == 'on'
        prefs.save()

        messages.success(request, 'Profile updated successfully.')
        return redirect('profile')

    return render(request, 'profile_edit.html', {'user': user, 'prefs': prefs})


@login_required(login_url='login')
def profile_kyc(request):
    user = request.user
    if request.method == 'POST':
        if user.national_id and user.kra_pin:
            user.kyc_status = 'pending'
            user.save()
            messages.success(request, 'KYC submitted for review.')
        else:
            messages.error(request, 'Please fill in National ID and KRA PIN in your profile first.')
    return redirect('profile')


# ─── Activity Logger ──────────────────────────────────────────────────────────

def log_activity(user, event_type, chama=None, amount=None, note=None):
    MemberActivity.objects.create(
        user=user,
        chama=chama,
        event_type=event_type,
        amount=amount,
        note=note,
    )


# ─── Chama ────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def chama_create(request):
    contribution_days = [1, 5, 10, 15, 20, 25, 28]
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Chama name is required.')
            return render(request, 'chama_create.html', {
                'contribution_days': contribution_days
            })
        chama = Chama.objects.create(
            name=name,
            description=request.POST.get('description', ''),
            created_by=request.user,
            contribution_amount=request.POST.get('contribution_amount', 0) or 0,
            contribution_day=request.POST.get('contribution_day', 5) or 5,
            meeting_day=request.POST.get('meeting_day', ''),
            currency=request.POST.get('currency', 'KES'),
        )
        Membership.objects.create(chama=chama, user=request.user, role='admin')
        request.session['active_chama_id'] = chama.id
        log_activity(request.user, 'chama_created', chama=chama)
        messages.success(request, f'"{chama.name}" created! You are the Admin.')
        return redirect('chama_members', chama_id=chama.id)
    return render(request, 'chama_create.html', {
        'contribution_days': contribution_days
    })


@login_required(login_url='login')
def chama_members(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    membership = get_object_or_404(Membership, chama=chama, user=request.user, active=True)
    memberships = chama.memberships.filter(active=True).select_related('user').order_by('joined_at')
    return render(request, 'chama_members.html', {
        'chama': chama, 'memberships': memberships,
        'my_membership': membership,
        'can_manage': membership.role in ('admin', 'treasurer'),
    })


@login_required(login_url='login')
def member_invite(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my = get_object_or_404(Membership, chama=chama, user=request.user, active=True)
    if my.role not in ('admin', 'treasurer'):
        messages.error(request, 'Only admins can invite members.')
        return redirect('chama_members', chama_id=chama_id)
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        role  = request.POST.get('role', 'member')
        try:
            invitee = User.objects.get(email=email)
        except User.DoesNotExist:
            messages.error(request, f'No account found for {email}.')
            return redirect('chama_members', chama_id=chama_id)
        if Membership.objects.filter(chama=chama, user=invitee).exists():
            messages.warning(request, f'{invitee.get_full_name()} is already a member.')
            return redirect('chama_members', chama_id=chama_id)
        Membership.objects.create(chama=chama, user=invitee, role=role)
        log_activity(invitee, 'chama_joined', chama=chama)
        messages.success(request, f'{invitee.get_full_name()} added as {role}.')
    return redirect('chama_members', chama_id=chama_id)


@login_required(login_url='login')
def member_role_update(request, chama_id, membership_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my = get_object_or_404(Membership, chama=chama, user=request.user, active=True)
    if my.role != 'admin':
        messages.error(request, 'Only admins can change roles.')
        return redirect('chama_members', chama_id=chama_id)
    if request.method == 'POST':
        target = get_object_or_404(Membership, id=membership_id, chama=chama)
        target.role = request.POST.get('role', 'member')
        target.save()
        messages.success(request, f'{target.user.get_full_name()} is now {target.role}.')
    return redirect('chama_members', chama_id=chama_id)


@login_required(login_url='login')
def member_remove(request, chama_id, membership_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my = get_object_or_404(Membership, chama=chama, user=request.user, active=True)
    if my.role != 'admin':
        messages.error(request, 'Only admins can remove members.')
        return redirect('chama_members', chama_id=chama_id)
    target = get_object_or_404(Membership, id=membership_id, chama=chama)
    if target.user == request.user:
        messages.error(request, "You can't remove yourself.")
        return redirect('chama_members', chama_id=chama_id)
    target.active = False
    target.save()
    messages.success(request, f'{target.user.get_full_name()} removed.')
    return redirect('chama_members', chama_id=chama_id)


# ─── Chama Settings (NEW) ─────────────────────────────────────────────────────

@login_required(login_url='login')
def chama_settings(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my    = get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    # Only admin can access settings
    if my.role != 'admin':
        messages.error(request, 'Only admins can access chama settings.')
        return redirect('dashboard')

    members           = Membership.objects.filter(chama=chama, active=True).select_related('user')
    contribution_days = [1, 5, 10, 15, 20, 25, 28]

    if request.method == 'POST':
        section = request.POST.get('section', '')

        # ── Group Identity ──
        if section == 'chama':
            name = request.POST.get('name', '').strip()
            if not name:
                messages.error(request, 'Chama name cannot be empty.')
                return redirect('chama_settings', chama_id=chama_id)
            chama.name        = name
            chama.description = request.POST.get('description', '').strip()
            chama.save()
            messages.success(request, 'Group identity updated successfully.')

        # ── Schedule & Contributions ──
        elif section == 'schedule':
            chama.meeting_day      = request.POST.get('meeting_day', chama.meeting_day)
            contrib_amount         = request.POST.get('contribution_amount', '').strip()
            chama.contribution_amount = Decimal(contrib_amount) if contrib_amount else chama.contribution_amount
            chama.contribution_day = int(request.POST.get('contribution_day', chama.contribution_day))
            max_members            = request.POST.get('max_members', '').strip()
            if max_members:
                chama.max_members = int(max_members)
            late_penalty = request.POST.get('late_penalty', '').strip()
            if late_penalty:
                chama.late_penalty = Decimal(late_penalty)
            loan_interest = request.POST.get('loan_interest', '').strip()
            if loan_interest:
                chama.loan_interest = Decimal(loan_interest)
            chama.save()
            messages.success(request, 'Schedule & contributions updated.')

        # ── Currency ──
        elif section == 'currency':
            currency = request.POST.get('currency', 'KES')
            if currency in ('KES', 'UGX', 'TZS', 'USD'):
                chama.currency = currency
                chama.save()
                messages.success(request, f'Currency updated to {currency}.')
            else:
                messages.error(request, 'Invalid currency selected.')

        # ── Member Role ──
        elif section == 'role':
            membership_id = request.POST.get('membership_id')
            new_role      = request.POST.get('role', 'member')
            if new_role not in ('admin', 'treasurer', 'secretary', 'member'):
                messages.error(request, 'Invalid role.')
                return redirect('chama_settings', chama_id=chama_id)
            target = get_object_or_404(Membership, id=membership_id, chama=chama, active=True)
            # Prevent demoting self if only admin
            if target.user == request.user and new_role != 'admin':
                admin_count = Membership.objects.filter(chama=chama, role='admin', active=True).count()
                if admin_count <= 1:
                    messages.error(request, "You're the only admin. Promote another member first.")
                    return redirect('chama_settings', chama_id=chama_id)
            target.role = new_role
            target.save()
            messages.success(request, f'{target.user.get_full_name()} is now {target.get_role_display()}.')

        # ── Invite Member ──
        elif section == 'invite':
            email = request.POST.get('invite_email', '').strip()
            if not email:
                messages.error(request, 'Email address is required.')
                return redirect('chama_settings', chama_id=chama_id)
            try:
                invitee = User.objects.get(email=email)
                if Membership.objects.filter(chama=chama, user=invitee, active=True).exists():
                    messages.warning(request, f'{invitee.get_full_name()} is already a member.')
                else:
                    Membership.objects.create(chama=chama, user=invitee, role='member')
                    log_activity(invitee, 'chama_joined', chama=chama)
                    messages.success(request, f'{invitee.get_full_name()} added successfully.')
            except User.DoesNotExist:
                messages.error(request, f'No ChamaPro account found for {email}.')

        return redirect('chama_settings', chama_id=chama_id)

    return render(request, 'settings.html', {
        'chama':             chama,
        'my_membership':     my,
        'members':           members,
        'contribution_days': contribution_days,
    })


@login_required(login_url='login')
def chama_delete(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my    = get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    if my.role != 'admin':
        messages.error(request, 'Only admins can delete a chama.')
        return redirect('chama_settings', chama_id=chama_id)

    if request.method == 'POST':
        name = chama.name
        # Clear session if deleted chama was active
        if request.session.get('active_chama_id') == chama_id:
            try:
                del request.session['active_chama_id']
            except KeyError:
                pass
        chama.delete()
        messages.success(request, f'"{name}" has been permanently deleted.')
        return redirect('dashboard')

    return redirect('chama_settings', chama_id=chama_id)


# ─── Contributions ────────────────────────────────────────────────────────────

@login_required(login_url='login')
def contributions(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my = get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    contribs = chama.contributions.select_related('member', 'recorded_by').order_by('-date', '-created_at')

    member_filter = request.GET.get('member', '')
    method_filter = request.GET.get('method', '')
    status_filter = request.GET.get('status', '')

    if member_filter:
        contribs = contribs.filter(member_id=member_filter)
    if method_filter:
        contribs = contribs.filter(payment_method=method_filter)
    if status_filter:
        contribs = contribs.filter(status=status_filter)

    today = datetime.date.today()
    total_confirmed  = chama.contributions.filter(status='confirmed').aggregate(t=Sum('amount'))['t'] or 0
    total_pending    = chama.contributions.filter(status='pending').aggregate(t=Sum('amount'))['t'] or 0
    total_this_month = chama.contributions.filter(
        status='confirmed', date__month=today.month, date__year=today.year,
    ).aggregate(t=Sum('amount'))['t'] or 0

    members = chama.memberships.filter(active=True).select_related('user')

    context = {
        'chama': chama,
        'my_membership': my,
        'can_manage': my.role in ('admin', 'treasurer'),
        'contributions': contribs,
        'members': members,
        'total_confirmed': total_confirmed,
        'total_pending': total_pending,
        'total_this_month': total_this_month,
        'member_filter': member_filter,
        'method_filter': method_filter,
        'status_filter': status_filter,
        'payment_methods': Contribution.PAYMENT_METHODS,
    }
    return render(request, 'contributions.html', context)


@login_required(login_url='login')
def contribution_add(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my = get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    if my.role not in ('admin', 'treasurer'):
        messages.error(request, 'Only admins and treasurers can record contributions.')
        return redirect('contributions', chama_id=chama_id)

    if request.method == 'POST':
        member_id = request.POST.get('member_id')
        amount    = request.POST.get('amount', '').strip()
        method    = request.POST.get('payment_method', 'cash')
        reference = request.POST.get('reference', '').strip()
        notes     = request.POST.get('notes', '').strip()
        date      = request.POST.get('date', '').strip()
        status    = request.POST.get('status', 'confirmed')

        errors = []
        if not member_id: errors.append('Please select a member.')
        if not amount:    errors.append('Amount is required.')
        try:
            amount = float(amount)
            if amount <= 0: errors.append('Amount must be greater than 0.')
        except (ValueError, TypeError):
            errors.append('Invalid amount.')

        if errors:
            for e in errors: messages.error(request, e)
            return redirect('contributions', chama_id=chama_id)

        member = get_object_or_404(User, id=member_id)
        Contribution.objects.create(
            chama=chama, member=member, amount=amount,
            payment_method=method, reference=reference or None,
            notes=notes or None, status=status,
            date=date or datetime.date.today(), recorded_by=request.user,
        )
        log_activity(member, 'contribution', chama=chama, amount=amount)
        messages.success(request, f'Contribution of KES {amount:,.2f} recorded for {member.get_full_name()}.')
        return redirect('contributions', chama_id=chama_id)

    return redirect('contributions', chama_id=chama_id)


@login_required(login_url='login')
def contribution_delete(request, chama_id, contribution_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my = get_object_or_404(Membership, chama=chama, user=request.user, active=True)
    if my.role not in ('admin', 'treasurer'):
        messages.error(request, 'Only admins and treasurers can delete contributions.')
        return redirect('contributions', chama_id=chama_id)
    contrib = get_object_or_404(Contribution, id=contribution_id, chama=chama)
    contrib.delete()
    messages.success(request, 'Contribution deleted.')
    return redirect('contributions', chama_id=chama_id)


@login_required(login_url='login')
def contribution_status_update(request, chama_id, contribution_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my = get_object_or_404(Membership, chama=chama, user=request.user, active=True)
    if my.role not in ('admin', 'treasurer'):
        messages.error(request, 'Permission denied.')
        return redirect('contributions', chama_id=chama_id)
    contrib = get_object_or_404(Contribution, id=contribution_id, chama=chama)
    new_status = request.POST.get('status')
    if new_status in ('confirmed', 'rejected', 'pending'):
        contrib.status = new_status
        contrib.save()
        messages.success(request, f'Contribution marked as {new_status}.')
    return redirect('contributions', chama_id=chama_id)


# ─── Penalties ────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def penalty_add(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my = get_object_or_404(Membership, chama=chama, user=request.user, active=True)
    if my.role not in ('admin', 'treasurer'):
        messages.error(request, 'Only admins and treasurers can issue penalties.')
        return redirect('contributions', chama_id=chama_id)
    if request.method == 'POST':
        member_id   = request.POST.get('member_id')
        amount      = request.POST.get('amount', '').strip()
        reason      = request.POST.get('reason', 'late_contribution')
        description = request.POST.get('description', '').strip()
        member = get_object_or_404(User, id=member_id)
        Penalty.objects.create(
            chama=chama, member=member, amount=amount, reason=reason,
            description=description or None, issued_by=request.user,
        )
        messages.success(request, f'Penalty of KES {amount} issued to {member.get_full_name()}.')
    return redirect('contributions', chama_id=chama_id)


# ─── Fines ────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def fines(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my    = get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    qs = chama.penalties.select_related('member', 'issued_by').order_by('-issued_at')

    status_filter = request.GET.get('status', '')
    member_filter = request.GET.get('member', '')
    reason_filter = request.GET.get('reason', '')

    if status_filter: qs = qs.filter(status=status_filter)
    if member_filter: qs = qs.filter(member_id=member_filter)
    if reason_filter: qs = qs.filter(reason=reason_filter)

    total_fines  = chama.penalties.aggregate(t=Sum('amount'))['t'] or Decimal('0')
    total_unpaid = chama.penalties.filter(status='unpaid').aggregate(t=Sum('amount'))['t'] or Decimal('0')
    total_paid   = chama.penalties.filter(status='paid').aggregate(t=Sum('amount'))['t'] or Decimal('0')
    total_waived = chama.penalties.filter(status='waived').aggregate(t=Sum('amount'))['t'] or Decimal('0')
    count_unpaid = chama.penalties.filter(status='unpaid').count()

    members = chama.memberships.filter(active=True).select_related('user')

    context = {
        'chama':           chama,
        'my_membership':   my,
        'can_manage':      my.role in ('admin', 'treasurer'),
        'penalties':       qs,
        'members':         members,
        'total_fines':     total_fines,
        'total_unpaid':    total_unpaid,
        'total_paid':      total_paid,
        'total_waived':    total_waived,
        'count_unpaid':    count_unpaid,
        'status_filter':   status_filter,
        'member_filter':   member_filter,
        'reason_filter':   reason_filter,
        'penalty_reasons': Penalty.PENALTY_REASONS,
        'status_choices':  Penalty.STATUS_CHOICES,
    }
    return render(request, 'chamapro/fines.html', context)


@login_required(login_url='login')
def fine_add(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my    = get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    if my.role not in ('admin', 'treasurer'):
        messages.error(request, 'Only admins and treasurers can issue fines.')
        return redirect('fines', chama_id=chama_id)

    if request.method == 'POST':
        member_id   = request.POST.get('member_id')
        amount      = request.POST.get('amount', '').strip()
        reason      = request.POST.get('reason', 'other')
        description = request.POST.get('description', '').strip()

        errors = []
        if not member_id: errors.append('Please select a member.')
        if not amount:    errors.append('Amount is required.')
        try:
            amount_val = Decimal(amount)
            if amount_val <= 0: errors.append('Amount must be greater than 0.')
        except Exception:
            errors.append('Invalid amount.')

        if errors:
            for e in errors: messages.error(request, e)
            return redirect('fines', chama_id=chama_id)

        member  = get_object_or_404(User, id=member_id)
        penalty = Penalty.objects.create(
            chama=chama, member=member,
            amount=Decimal(amount), reason=reason,
            description=description or None,
            issued_by=request.user, status='unpaid',
        )
        log_activity(member, 'fine_issued', chama=chama, amount=float(penalty.amount),
                     note=f'{penalty.get_reason_display()} — KES {penalty.amount}')
        messages.success(request, f'Fine of KES {penalty.amount:,.2f} issued to {member.get_full_name()}.')

    return redirect('fines', chama_id=chama_id)


@login_required(login_url='login')
def fine_update_status(request, chama_id, penalty_id):
    chama   = get_object_or_404(Chama, id=chama_id)
    my      = get_object_or_404(Membership, chama=chama, user=request.user, active=True)
    penalty = get_object_or_404(Penalty, id=penalty_id, chama=chama)

    if my.role not in ('admin', 'treasurer'):
        messages.error(request, 'Permission denied.')
        return redirect('fines', chama_id=chama_id)

    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in ('unpaid', 'paid', 'waived'):
            old_status     = penalty.status
            penalty.status = new_status
            penalty.save()
            if new_status == 'paid' and old_status != 'paid':
                log_activity(penalty.member, 'fine_paid', chama=chama,
                             amount=float(penalty.amount),
                             note=f'Fine paid — KES {penalty.amount}')
                messages.success(request, f'Fine marked as paid for {penalty.member.get_full_name()}.')
            elif new_status == 'waived':
                messages.success(request, f'Fine waived for {penalty.member.get_full_name()}.')
            else:
                messages.success(request, 'Fine status updated.')

    return redirect('fines', chama_id=chama_id)


@login_required(login_url='login')
def fine_delete(request, chama_id, penalty_id):
    chama   = get_object_or_404(Chama, id=chama_id)
    my      = get_object_or_404(Membership, chama=chama, user=request.user, active=True)
    penalty = get_object_or_404(Penalty, id=penalty_id, chama=chama)

    if my.role != 'admin':
        messages.error(request, 'Only admins can delete fines.')
        return redirect('fines', chama_id=chama_id)

    penalty.delete()
    messages.success(request, 'Fine deleted.')
    return redirect('fines', chama_id=chama_id)


@login_required(login_url='login')
def fine_pay_mpesa(request, chama_id, penalty_id):
    chama   = get_object_or_404(Chama, id=chama_id)
    my      = get_object_or_404(Membership, chama=chama, user=request.user, active=True)
    penalty = get_object_or_404(Penalty, id=penalty_id, chama=chama, status='unpaid')

    if request.user != penalty.member and my.role not in ('admin', 'treasurer'):
        return JsonResponse({'success': False, 'error': 'Permission denied.'})

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        data  = json.loads(request.body)
        phone = data.get('phone', '').strip()
    except Exception:
        phone = request.POST.get('phone', '').strip()

    if not phone:
        phone = penalty.member.mpesa_number or ''

    if not phone:
        return JsonResponse({'success': False, 'error': 'No phone number provided.'})

    try:
        result = mpesa.stk_push(
            phone_number=phone,
            amount=int(penalty.amount),
            account_reference=f'FINE-{penalty.id}',
            transaction_desc=f'Fine: {penalty.get_reason_display()}',
        )
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

    if result.get('ResponseCode') == '0':
        return JsonResponse({
            'success': True,
            'message': f'STK Push sent to {phone}. Enter your M-Pesa PIN.',
            'checkout_request_id': result.get('CheckoutRequestID'),
        })
    return JsonResponse({'success': False, 'error': result.get('errorMessage', 'STK Push failed.')})


@login_required(login_url='login')
def fine_report(request, chama_id):
    chama       = get_object_or_404(Chama, id=chama_id)
    my          = get_object_or_404(Membership, chama=chama, user=request.user, active=True)
    memberships = chama.memberships.filter(active=True).select_related('user')
    report_data = []

    for m in memberships:
        mf     = chama.penalties.filter(member=m.user)
        total  = mf.aggregate(t=Sum('amount'))['t'] or Decimal('0')
        unpaid = mf.filter(status='unpaid').aggregate(t=Sum('amount'))['t'] or Decimal('0')
        paid   = mf.filter(status='paid').aggregate(t=Sum('amount'))['t'] or Decimal('0')
        waived = mf.filter(status='waived').aggregate(t=Sum('amount'))['t'] or Decimal('0')
        latest = mf.order_by('-issued_at').first()
        report_data.append({
            'member': m.user, 'role': m.get_role_display(),
            'total': total, 'unpaid': unpaid, 'paid': paid,
            'waived': waived, 'count': mf.count(), 'latest_fine': latest,
        })

    report_data.sort(key=lambda x: x['unpaid'], reverse=True)

    return render(request, 'chamapro/fine_report.html', {
        'chama':         chama,
        'my_membership': my,
        'can_manage':    my.role in ('admin', 'treasurer'),
        'report_data':   report_data,
    })


# ─── Loans ────────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def loans(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my = get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    all_loans = chama.loans.select_related('member', 'approved_by').order_by('-created_at')

    status_filter = request.GET.get('status', '')
    member_filter = request.GET.get('member', '')
    if status_filter:
        all_loans = all_loans.filter(status=status_filter)
    if member_filter:
        all_loans = all_loans.filter(member_id=member_filter)

    total_disbursed   = chama.loans.filter(status__in=['active', 'overdue', 'repaid']).aggregate(t=Sum('amount'))['t'] or Decimal('0')
    total_outstanding = chama.loans.filter(status__in=['active', 'overdue']).aggregate(t=Sum('amount'))['t'] or Decimal('0')
    total_pending     = chama.loans.filter(status='pending').count()
    total_overdue     = chama.loans.filter(status='overdue').count()

    members = chama.memberships.filter(active=True).select_related('user')

    context = {
        'chama': chama,
        'my_membership': my,
        'can_manage': my.role in ('admin', 'treasurer'),
        'loans': all_loans,
        'members': members,
        'total_disbursed': total_disbursed,
        'total_outstanding': total_outstanding,
        'total_pending': total_pending,
        'total_overdue': total_overdue,
        'status_filter': status_filter,
        'member_filter': member_filter,
        'loan_statuses': Loan.STATUS_CHOICES,
        'purpose_choices': Loan.PURPOSE_CHOICES,
    }
    return render(request, 'loans.html', context)


@login_required(login_url='login')
def loan_apply(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my = get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    if request.method == 'POST':
        member_id    = request.POST.get('member_id') or request.user.id
        amount       = request.POST.get('amount', '').strip()
        interest     = request.POST.get('interest_rate', '10').strip()
        term         = request.POST.get('term_months', '3').strip()
        purpose      = request.POST.get('purpose', 'other')
        description  = request.POST.get('description', '').strip()
        applied_date = request.POST.get('applied_at', '').strip()

        errors = []
        if not amount: errors.append('Loan amount is required.')
        try:
            amount_val = float(amount)
            if amount_val <= 0: errors.append('Amount must be greater than 0.')
        except (ValueError, TypeError):
            errors.append('Invalid amount.')

        if errors:
            for e in errors: messages.error(request, e)
            return redirect('loans', chama_id=chama_id)

        if my.role in ('admin', 'treasurer') and member_id:
            member = get_object_or_404(User, id=member_id)
        else:
            member = request.user

        Loan.objects.create(
            chama=chama, member=member,
            amount=Decimal(amount),
            interest_rate=Decimal(interest or '10'),
            term_months=int(term or 3),
            purpose=purpose, description=description or None,
            status='pending',
            applied_at=applied_date or datetime.date.today(),
        )
        messages.success(request, f'Loan application of KES {amount_val:,.2f} submitted for {member.get_full_name()}.')
        return redirect('loans', chama_id=chama_id)

    return redirect('loans', chama_id=chama_id)


@login_required(login_url='login')
def loan_approve(request, chama_id, loan_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my = get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    if my.role not in ('admin', 'treasurer'):
        messages.error(request, 'Only admins and treasurers can approve loans.')
        return redirect('loans', chama_id=chama_id)

    loan = get_object_or_404(Loan, id=loan_id, chama=chama)

    if request.method == 'POST':
        action = request.POST.get('action')
        from dateutil.relativedelta import relativedelta

        if action == 'approve' and loan.status == 'pending':
            loan.status      = 'active'
            loan.approved_by = request.user
            loan.approved_at = datetime.date.today()
            loan.due_date    = datetime.date.today() + relativedelta(months=loan.term_months)
            loan.save()
            log_activity(loan.member, 'loan_approved', chama=chama, amount=float(loan.amount))
            messages.success(request, f'Loan of KES {loan.amount:,.2f} approved for {loan.member.get_full_name()}.')
        elif action == 'reject' and loan.status == 'pending':
            loan.status = 'rejected'
            loan.save()
            messages.warning(request, f'Loan application rejected for {loan.member.get_full_name()}.')

    return redirect('loans', chama_id=chama_id)


@login_required(login_url='login')
def loan_repayment_add(request, chama_id, loan_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my = get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    if my.role not in ('admin', 'treasurer'):
        messages.error(request, 'Only admins and treasurers can record repayments.')
        return redirect('loans', chama_id=chama_id)

    loan = get_object_or_404(Loan, id=loan_id, chama=chama)

    if request.method == 'POST':
        amount    = request.POST.get('amount', '').strip()
        method    = request.POST.get('payment_method', 'cash')
        reference = request.POST.get('reference', '').strip()
        date      = request.POST.get('date', '').strip()

        try:
            amount_val = Decimal(amount)
        except Exception:
            messages.error(request, 'Invalid amount.')
            return redirect('loans', chama_id=chama_id)

        LoanRepayment.objects.create(
            loan=loan, amount=amount_val, payment_method=method,
            reference=reference or None,
            date=date or datetime.date.today(),
            recorded_by=request.user, status='confirmed',
        )

        if loan.balance() <= 0:
            loan.status = 'repaid'
            loan.save()
            log_activity(loan.member, 'loan_repaid', chama=chama, amount=float(loan.amount))
            messages.success(request, f'Loan fully repaid by {loan.member.get_full_name()}! 🎉')
        else:
            messages.success(request, f'Repayment of KES {amount_val:,.2f} recorded. Balance: KES {loan.balance():,.2f}')

    return redirect('loans', chama_id=chama_id)


@login_required(login_url='login')
def loan_detail(request, chama_id, loan_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my = get_object_or_404(Membership, chama=chama, user=request.user, active=True)
    loan = get_object_or_404(Loan, id=loan_id, chama=chama)
    repayments = loan.repayments.all()

    context = {
        'chama': chama,
        'my_membership': my,
        'can_manage': my.role in ('admin', 'treasurer'),
        'loan': loan,
        'repayments': repayments,
        'payment_methods': LoanRepayment.PAYMENT_METHODS,
    }
    return render(request, 'loan_detail.html', context)


# ─── Reports ──────────────────────────────────────────────────────────────────

@login_required(login_url='login')
def reports(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my = get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    today = datetime.date.today()

    total_contributions = chama.contributions.filter(
        status='confirmed'
    ).aggregate(t=Sum('amount'))['t'] or Decimal('0')

    this_month_contributions = chama.contributions.filter(
        status='confirmed', date__month=today.month, date__year=today.year
    ).aggregate(t=Sum('amount'))['t'] or Decimal('0')

    monthly_data = []
    for i in range(11, -1, -1):
        d = (today.replace(day=1) - datetime.timedelta(days=i * 30)).replace(day=1)
        month_total = chama.contributions.filter(
            status='confirmed', date__month=d.month, date__year=d.year
        ).aggregate(t=Sum('amount'))['t'] or 0
        monthly_data.append({'label': d.strftime('%b %Y'), 'amount': float(month_total)})

    memberships = chama.memberships.filter(active=True).select_related('user')
    member_summary = []
    for m in memberships:
        paid = chama.contributions.filter(
            member=m.user, status='confirmed'
        ).aggregate(t=Sum('amount'))['t'] or Decimal('0')
        last_contrib = chama.contributions.filter(
            member=m.user, status='confirmed'
        ).order_by('-date').first()
        member_summary.append({
            'member': m.user,
            'role': m.get_role_display(),
            'total_paid': paid,
            'last_contribution': last_contrib.date if last_contrib else None,
            'arrears': max((chama.contribution_amount or Decimal('0')) - paid, Decimal('0')),
        })

    total_loans        = chama.loans.filter(status__in=['active', 'overdue', 'repaid']).aggregate(t=Sum('amount'))['t'] or Decimal('0')
    total_outstanding  = chama.loans.filter(status__in=['active', 'overdue']).aggregate(t=Sum('amount'))['t'] or Decimal('0')
    active_loans       = chama.loans.filter(status__in=['active', 'overdue']).count()
    total_overdue      = chama.loans.filter(status='overdue').count()
    total_pending      = chama.loans.filter(status='pending').count()
    total_repaid_count = chama.loans.filter(status='repaid').count()

    total_interest_earned = Decimal('0')
    for loan in chama.loans.filter(status='repaid'):
        total_interest_earned += loan.interest_amount()

    total_penalties  = chama.penalties.aggregate(t=Sum('amount'))['t'] or Decimal('0')
    unpaid_penalties = chama.penalties.filter(status='unpaid').aggregate(t=Sum('amount'))['t'] or Decimal('0')

    contrib_to_loan_ratio = 0
    if total_contributions > 0:
        contrib_to_loan_ratio = min(int((total_outstanding / total_contributions) * 100), 100)

    repayment_rate = 0
    if total_loans > 0:
        repaid_amount = chama.loans.filter(status='repaid').aggregate(t=Sum('amount'))['t'] or Decimal('0')
        repayment_rate = int((repaid_amount / total_loans) * 100)

    context = {
        'chama': chama,
        'my_membership': my,
        'can_manage': my.role in ('admin', 'treasurer'),
        'total_contributions': total_contributions,
        'this_month_contributions': this_month_contributions,
        'monthly_labels': [d['label'] for d in monthly_data],
        'monthly_amounts': [d['amount'] for d in monthly_data],
        'member_summary': member_summary,
        'total_members': chama.member_count(),
        'total_loans': total_loans,
        'total_outstanding': total_outstanding,
        'total_interest_earned': total_interest_earned,
        'active_loans': active_loans,
        'total_overdue': total_overdue,
        'total_pending': total_pending,
        'total_repaid': total_repaid_count,
        'total_penalties': total_penalties,
        'unpaid_penalties': unpaid_penalties,
        'contrib_to_loan_ratio': contrib_to_loan_ratio,
        'repayment_rate': repayment_rate,
        'today': today,
    }
    return render(request, 'reports.html', context)


# ── Export: CSV ───────────────────────────────────────────────────────────────

@login_required(login_url='login')
def export_contributions_csv(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{chama.name}_contributions_{datetime.date.today()}.csv"'

    writer = csv.writer(response)
    writer.writerow(['#', 'Member', 'Email', 'Amount (KES)', 'Date', 'Payment Method', 'Reference', 'Status', 'Recorded By'])
    for i, c in enumerate(chama.contributions.select_related('member', 'recorded_by').order_by('-date'), 1):
        writer.writerow([
            i, c.member.get_full_name() or c.member.username, c.member.email,
            c.amount, c.date, c.get_payment_method_display(),
            c.reference or '', c.get_status_display(),
            c.recorded_by.get_full_name() if c.recorded_by else '',
        ])
    return response


@login_required(login_url='login')
def export_loans_csv(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{chama.name}_loans_{datetime.date.today()}.csv"'

    writer = csv.writer(response)
    writer.writerow(['#', 'Member', 'Amount', 'Interest Rate %', 'Interest', 'Total Payable', 'Repaid', 'Balance', 'Status', 'Purpose', 'Applied', 'Due Date'])
    for i, loan in enumerate(chama.loans.select_related('member').order_by('-created_at'), 1):
        writer.writerow([
            i, loan.member.get_full_name() or loan.member.username,
            loan.amount, loan.interest_rate, loan.interest_amount(),
            loan.total_payable(), loan.total_repaid(), loan.balance(),
            loan.get_status_display(), loan.get_purpose_display(),
            loan.applied_at, loan.due_date or '',
        ])
    return response


@login_required(login_url='login')
def export_members_csv(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{chama.name}_members_{datetime.date.today()}.csv"'

    writer = csv.writer(response)
    writer.writerow(['#', 'Name', 'Email', 'Phone', 'Role', 'Joined', 'Total Contributed (KES)', 'Arrears (KES)'])
    for i, m in enumerate(chama.memberships.filter(active=True).select_related('user').order_by('joined_at'), 1):
        paid = chama.contributions.filter(member=m.user, status='confirmed').aggregate(t=Sum('amount'))['t'] or Decimal('0')
        arrears = max((chama.contribution_amount or Decimal('0')) - paid, Decimal('0'))
        writer.writerow([
            i, m.user.get_full_name() or m.user.username, m.user.email,
            m.user.phone or '', m.get_role_display(), m.joined_at.date(),
            paid, arrears,
        ])
    return response


# ── Export: PDF ───────────────────────────────────────────────────────────────

@login_required(login_url='login')
def export_report_pdf(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=18*mm, rightMargin=18*mm,
                            topMargin=16*mm, bottomMargin=16*mm)

    styles = getSampleStyleSheet()
    brand       = colors.HexColor('#0d6e4f')
    light_brand = colors.HexColor('#e6f5f0')
    story = []

    title_style = ParagraphStyle('title', fontSize=20, fontName='Helvetica-Bold', textColor=brand, spaceAfter=2)
    sub_style   = ParagraphStyle('sub', fontSize=10, textColor=colors.HexColor('#6b7280'), spaceAfter=12)
    h2_style    = ParagraphStyle('h2', fontSize=13, fontName='Helvetica-Bold', textColor=brand, spaceBefore=14, spaceAfter=6)
    normal      = ParagraphStyle('norm', fontSize=9, leading=14)

    story.append(Paragraph(f'{chama.name}', title_style))
    story.append(Paragraph(f'Financial Report  ·  Generated {datetime.date.today().strftime("%B %d, %Y")}', sub_style))
    story.append(HRFlowable(width='100%', thickness=1, color=brand))
    story.append(Spacer(1, 8))

    total_contribs  = chama.contributions.filter(status='confirmed').aggregate(t=Sum('amount'))['t'] or Decimal('0')
    total_loans_amt = chama.loans.filter(status__in=['active', 'overdue', 'repaid']).aggregate(t=Sum('amount'))['t'] or Decimal('0')
    outstanding     = chama.loans.filter(status__in=['active', 'overdue']).aggregate(t=Sum('amount'))['t'] or Decimal('0')
    total_penalties = chama.penalties.aggregate(t=Sum('amount'))['t'] or Decimal('0')

    summary_data = [
        ['Metric', 'Value'],
        ['Total Contributions (Confirmed)', f'KES {total_contribs:,.2f}'],
        ['Total Loans Disbursed', f'KES {total_loans_amt:,.2f}'],
        ['Outstanding Loans', f'KES {outstanding:,.2f}'],
        ['Total Penalties/Fines', f'KES {total_penalties:,.2f}'],
        ['Active Members', str(chama.member_count())],
    ]
    st = Table(summary_data, colWidths=[100*mm, 60*mm])
    st.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), brand), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_brand]),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#e5e7eb')),
        ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(Paragraph('Financial Summary', h2_style))
    story.append(st)
    story.append(Spacer(1, 10))

    story.append(Paragraph('Member Contribution Summary', h2_style))
    rows = [['#', 'Member', 'Email', 'Total Paid (KES)', 'Last Contribution']]
    for i, m in enumerate(chama.memberships.filter(active=True).select_related('user').order_by('joined_at'), 1):
        paid = chama.contributions.filter(member=m.user, status='confirmed').aggregate(t=Sum('amount'))['t'] or Decimal('0')
        last = chama.contributions.filter(member=m.user, status='confirmed').order_by('-date').first()
        rows.append([str(i), m.user.get_full_name() or m.user.username, m.user.email,
                     f'{paid:,.2f}', last.date.strftime('%b %d, %Y') if last else '—'])
    ct = Table(rows, colWidths=[10*mm, 50*mm, 55*mm, 35*mm, 30*mm])
    ct.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), brand), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_brand]),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#e5e7eb')),
        ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(ct)
    story.append(Spacer(1, 10))

    story.append(Paragraph('Loan Summary', h2_style))
    loan_rows = [['#', 'Member', 'Amount', 'Interest', 'Total Payable', 'Repaid', 'Balance', 'Status']]
    for i, loan in enumerate(chama.loans.select_related('member').order_by('-created_at'), 1):
        loan_rows.append([str(i), loan.member.get_full_name() or loan.member.username,
                          f'{loan.amount:,.2f}', f'{loan.interest_rate}%',
                          f'{loan.total_payable():,.2f}', f'{loan.total_repaid():,.2f}',
                          f'{loan.balance():,.2f}', loan.get_status_display()])
    if len(loan_rows) > 1:
        lt = Table(loan_rows, colWidths=[8*mm, 38*mm, 22*mm, 16*mm, 24*mm, 20*mm, 20*mm, 22*mm])
        lt.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), brand), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,-1), 7.5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_brand]),
            ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#e5e7eb')),
            ('TOPPADDING', (0,0), (-1,-1), 4), ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('LEFTPADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(lt)
    else:
        story.append(Paragraph('No loans recorded yet.', normal))

    story.append(Spacer(1, 16))
    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#e5e7eb')))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f'Generated by ChamaPro  ·  {chama.name}  ·  {datetime.datetime.now().strftime("%B %d, %Y %H:%M")}',
        ParagraphStyle('footer', fontSize=8, textColor=colors.HexColor('#9ca3af'), alignment=TA_CENTER)
    ))

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{chama.name}_report_{datetime.date.today()}.pdf"'
    return response


# ── Export: Excel ─────────────────────────────────────────────────────────────

@login_required(login_url='login')
def export_report_excel(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Run: pip install openpyxl')
        return redirect('reports', chama_id=chama_id)

    wb = openpyxl.Workbook()
    brand_fill  = PatternFill('solid', fgColor='0D6E4F')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    alt_fill    = PatternFill('solid', fgColor='E6F5F0')
    border_side = Side(style='thin', color='E5E7EB')
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    def style_header(ws, row, cols):
        for c in range(1, cols+1):
            cell = ws.cell(row=row, column=c)
            cell.fill = brand_fill; cell.font = header_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border

    def style_row(ws, row, cols, alt=False):
        for c in range(1, cols+1):
            cell = ws.cell(row=row, column=c)
            if alt: cell.fill = alt_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    ws1 = wb.active; ws1.title = 'Summary'
    ws1['A1'] = f'{chama.name} – Financial Report'
    ws1['A1'].font = Font(bold=True, size=14, color='0D6E4F')
    ws1['A2'] = f'Generated: {datetime.date.today().strftime("%B %d, %Y")}'
    ws1['A2'].font = Font(size=10, color='6B7280')
    for ci, h in enumerate(['Metric', 'Value'], 1):
        ws1.cell(row=4, column=ci, value=h)
    style_header(ws1, 4, 2)

    total_contribs  = chama.contributions.filter(status='confirmed').aggregate(t=Sum('amount'))['t'] or Decimal('0')
    total_loans_amt = chama.loans.filter(status__in=['active','overdue','repaid']).aggregate(t=Sum('amount'))['t'] or Decimal('0')
    outstanding     = chama.loans.filter(status__in=['active','overdue']).aggregate(t=Sum('amount'))['t'] or Decimal('0')
    total_penalties = chama.penalties.aggregate(t=Sum('amount'))['t'] or Decimal('0')

    for ri, (metric, val) in enumerate([
        ('Total Contributions', float(total_contribs)),
        ('Total Loans Disbursed', float(total_loans_amt)),
        ('Outstanding Loans', float(outstanding)),
        ('Total Penalties', float(total_penalties)),
        ('Active Members', chama.member_count()),
    ], 5):
        ws1.cell(row=ri, column=1, value=metric)
        ws1.cell(row=ri, column=2, value=val)
        style_row(ws1, ri, 2, alt=(ri % 2 == 0))
    ws1.column_dimensions['A'].width = 36
    ws1.column_dimensions['B'].width = 20

    ws2 = wb.create_sheet('Contributions')
    h2 = ['#', 'Member', 'Email', 'Amount (KES)', 'Date', 'Method', 'Reference', 'Status']
    for ci, h in enumerate(h2, 1): ws2.cell(row=1, column=ci, value=h)
    style_header(ws2, 1, len(h2))
    for ri, c in enumerate(chama.contributions.select_related('member').order_by('-date'), 2):
        vals = [ri-1, c.member.get_full_name() or c.member.username, c.member.email,
                float(c.amount), str(c.date), c.get_payment_method_display(),
                c.reference or '', c.get_status_display()]
        for ci, v in enumerate(vals, 1): ws2.cell(row=ri, column=ci, value=v)
        style_row(ws2, ri, len(h2), alt=(ri % 2 == 0))
    for ci, w in enumerate([6,26,28,16,14,14,18,14], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    ws3 = wb.create_sheet('Members')
    h3 = ['#', 'Name', 'Email', 'Phone', 'Role', 'Joined', 'Total Contributed', 'Arrears']
    for ci, h in enumerate(h3, 1): ws3.cell(row=1, column=ci, value=h)
    style_header(ws3, 1, len(h3))
    for ri, m in enumerate(chama.memberships.filter(active=True).select_related('user').order_by('joined_at'), 2):
        paid = chama.contributions.filter(member=m.user, status='confirmed').aggregate(t=Sum('amount'))['t'] or Decimal('0')
        arrears = max((chama.contribution_amount or Decimal('0')) - paid, Decimal('0'))
        vals = [ri-1, m.user.get_full_name() or m.user.username, m.user.email,
                m.user.phone or '', m.get_role_display(), str(m.joined_at.date()),
                float(paid), float(arrears)]
        for ci, v in enumerate(vals, 1): ws3.cell(row=ri, column=ci, value=v)
        style_row(ws3, ri, len(h3), alt=(ri % 2 == 0))
    for ci, w in enumerate([6,26,28,18,14,14,22,16], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    ws4 = wb.create_sheet('Loans')
    h4 = ['#', 'Member', 'Amount', 'Rate %', 'Interest', 'Total Payable', 'Repaid', 'Balance', 'Status', 'Purpose', 'Applied', 'Due Date']
    for ci, h in enumerate(h4, 1): ws4.cell(row=1, column=ci, value=h)
    style_header(ws4, 1, len(h4))
    for ri, loan in enumerate(chama.loans.select_related('member').order_by('-created_at'), 2):
        vals = [ri-1, loan.member.get_full_name() or loan.member.username,
                float(loan.amount), float(loan.interest_rate), float(loan.interest_amount()),
                float(loan.total_payable()), float(loan.total_repaid()), float(loan.balance()),
                loan.get_status_display(), loan.get_purpose_display(),
                str(loan.applied_at), str(loan.due_date) if loan.due_date else '']
        for ci, v in enumerate(vals, 1): ws4.cell(row=ri, column=ci, value=v)
        style_row(ws4, ri, len(h4), alt=(ri % 2 == 0))
    for ci, w in enumerate([6,24,16,10,14,16,14,14,16,18,14,14], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{chama.name}_report_{datetime.date.today()}.xlsx"'
    return response


# ─── M-Pesa Views ─────────────────────────────────────────────────────────────

@login_required(login_url='login')
def mpesa_stk_push(request, chama_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    chama = get_object_or_404(Chama, id=chama_id)
    get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    try:
        data    = json.loads(request.body)
        phone   = data.get('phone', '').strip()
        amount  = data.get('amount', 0)
        tx_type = data.get('type', 'contribution')
        loan_id = data.get('loan_id')
    except (json.JSONDecodeError, AttributeError):
        phone   = request.POST.get('phone', '').strip()
        amount  = request.POST.get('amount', 0)
        tx_type = request.POST.get('type', 'contribution')
        loan_id = request.POST.get('loan_id')

    if not phone or not amount:
        return JsonResponse({'success': False, 'error': 'Phone and amount are required.'})

    try:
        amount = int(float(amount))
        if amount < 1:
            raise ValueError
    except (ValueError, TypeError):
        return JsonResponse({'success': False, 'error': 'Invalid amount.'})

    account_ref = f'CP-{chama.id}'
    tx_desc     = 'Contribution' if tx_type == 'contribution' else 'Loan Repayment'

    try:
        response = mpesa.stk_push(
            phone_number=phone,
            amount=amount,
            account_reference=account_ref,
            transaction_desc=tx_desc,
        )
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'M-Pesa error: {str(e)}'})

    if response.get('ResponseCode') == '0':
        MpesaTransaction.objects.create(
            chama=chama,
            member=request.user,
            phone=phone,
            amount=amount,
            transaction_type=tx_type,
            checkout_request_id=response.get('CheckoutRequestID'),
            merchant_request_id=response.get('MerchantRequestID'),
            status='pending',
        )
        return JsonResponse({
            'success': True,
            'message': f'STK Push sent to {phone}. Check your phone and enter your M-Pesa PIN.',
            'checkout_request_id': response.get('CheckoutRequestID'),
        })
    else:
        return JsonResponse({
            'success': False,
            'error': response.get('errorMessage') or response.get('ResponseDescription', 'STK Push failed.'),
        })


@login_required(login_url='login')
def mpesa_stk_query(request, chama_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
        checkout_request_id = data.get('checkout_request_id')
    except Exception:
        checkout_request_id = request.POST.get('checkout_request_id')

    if not checkout_request_id:
        return JsonResponse({'success': False, 'error': 'Missing checkout_request_id'})

    try:
        tx = MpesaTransaction.objects.get(
            checkout_request_id=checkout_request_id,
            member=request.user,
        )
    except MpesaTransaction.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Transaction not found.'})

    if tx.status in ('success', 'failed', 'cancelled'):
        return JsonResponse({'success': True, 'status': tx.status, 'receipt': tx.mpesa_receipt})

    try:
        result = mpesa.stk_query(checkout_request_id)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

    result_code = str(result.get('ResultCode', ''))

    if result_code == '0':
        return JsonResponse({'success': True, 'status': 'success'})
    elif result_code in ('1032', '1001'):
        return JsonResponse({'success': True, 'status': 'cancelled'})
    elif result_code:
        return JsonResponse({'success': True, 'status': 'failed', 'error': result.get('ResultDesc', '')})
    else:
        return JsonResponse({'success': True, 'status': 'pending'})


@csrf_exempt
@require_POST
def mpesa_callback(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ResultCode': 1, 'ResultDesc': 'Invalid JSON'})

    try:
        body         = data['Body']['stkCallback']
        checkout_req = body.get('CheckoutRequestID')
        result_code  = str(body.get('ResultCode', ''))
        result_desc  = body.get('ResultDesc', '')
    except (KeyError, TypeError):
        return JsonResponse({'ResultCode': 1, 'ResultDesc': 'Malformed callback'})

    try:
        tx = MpesaTransaction.objects.get(checkout_request_id=checkout_req)
    except MpesaTransaction.DoesNotExist:
        return JsonResponse({'ResultCode': 0, 'ResultDesc': 'OK'})

    tx.result_code = result_code
    tx.result_desc = result_desc

    if result_code == '0':
        items = body.get('CallbackMetadata', {}).get('Item', [])
        meta  = {item['Name']: item.get('Value') for item in items}

        tx.mpesa_receipt = meta.get('MpesaReceiptNumber')
        tx.status        = 'success'
        tx.save()

        if tx.transaction_type == 'contribution' and not tx.contribution:
            contrib = Contribution.objects.create(
                chama=tx.chama,
                member=tx.member,
                amount=tx.amount,
                payment_method='mpesa',
                reference=tx.mpesa_receipt,
                notes='Auto-recorded via M-Pesa STK Push',
                status='confirmed',
                date=datetime.date.today(),
            )
            tx.contribution = contrib
            tx.save()

        elif tx.transaction_type == 'loan_repayment' and not tx.loan_repayment:
            loan = Loan.objects.filter(
                chama=tx.chama,
                member=tx.member,
                status__in=['active', 'overdue']
            ).first()
            if loan:
                repayment = LoanRepayment.objects.create(
                    loan=loan,
                    amount=tx.amount,
                    payment_method='mpesa',
                    reference=tx.mpesa_receipt,
                    status='confirmed',
                    date=datetime.date.today(),
                )
                tx.loan_repayment = repayment
                tx.save()
                if loan.balance() <= 0:
                    loan.status = 'repaid'
                    loan.save()
    else:
        tx.status = 'cancelled' if result_code == '1032' else 'failed'
        tx.save()

    return JsonResponse({'ResultCode': 0, 'ResultDesc': 'Accepted'})


@login_required(login_url='login')
def mpesa_transactions(request, chama_id):
    chama = get_object_or_404(Chama, id=chama_id)
    my    = get_object_or_404(Membership, chama=chama, user=request.user, active=True)

    transactions = MpesaTransaction.objects.filter(chama=chama).select_related('member').order_by('-created_at')

    context = {
        'chama': chama,
        'my_membership': my,
        'transactions': transactions,
        'can_manage': my.role in ('admin', 'treasurer'),
    }
    return render(request, 'mpesa_transactions.html', context)


# ─── Upgrade / Subscription ───────────────────────────────────────────────────

PLAN_PRICES = {
    'premium': {'monthly': 999,  'annual': 9588},
    'pro':     {'monthly': 2499, 'annual': 23988},
}


@login_required(login_url='login')
def upgrade(request):
    membership   = Membership.objects.filter(user=request.user).select_related('chama').first()
    active_chama = membership.chama if membership else None
    return render(request, 'chamapro/upgrade.html', {
        'active_chama': active_chama,
    })


@login_required(login_url='login')
def upgrade_pay(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    try:
        body    = json.loads(request.body)
        phone   = body.get('phone', '').strip()
        amount  = int(body.get('amount', 0))
        plan    = body.get('plan', '')
        billing = body.get('billing', 'monthly')
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

    if not phone or not phone.startswith('254') or len(phone) != 12:
        return JsonResponse({'success': False, 'error': 'Invalid phone number. Use 254XXXXXXXXX format.'})
    if plan not in PLAN_PRICES:
        return JsonResponse({'success': False, 'error': 'Invalid plan selected.'})
    if amount != PLAN_PRICES[plan][billing]:
        return JsonResponse({'success': False, 'error': 'Amount mismatch. Please refresh and try again.'})

    try:
        description = f'ChamaPro {plan.title()} Plan - {billing.title()}'
        result = mpesa.stk_push(
            phone_number=phone,
            amount=amount,
            account_reference='ChamaPro-Upgrade',
            transaction_desc=description,
        )

        if result.get('ResponseCode') == '0':
            checkout_id = result.get('CheckoutRequestID')
            from .models import SubscriptionPayment
            SubscriptionPayment.objects.create(
                user=request.user,
                plan=plan,
                billing_cycle=billing,
                amount=amount,
                phone=phone,
                checkout_request_id=checkout_id,
                status='pending',
            )
            return JsonResponse({
                'success': True,
                'message': f'STK Push sent to +{phone}. Enter your M-Pesa PIN to complete.',
                'checkout_request_id': checkout_id,
            })
        else:
            return JsonResponse({'success': False, 'error': result.get('errorMessage', 'M-Pesa request failed.')})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required(login_url='login')
def upgrade_poll(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error'}, status=405)

    try:
        body        = json.loads(request.body)
        checkout_id = body.get('checkout_request_id', '')
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error'}, status=400)

    try:
        from .models import SubscriptionPayment
        payment = SubscriptionPayment.objects.get(
            checkout_request_id=checkout_id,
            user=request.user,
        )

        if payment.status == 'completed':
            return JsonResponse({'status': 'success', 'receipt': payment.mpesa_receipt})
        elif payment.status == 'failed':
            return JsonResponse({'status': 'failed', 'error': payment.failure_reason or 'Payment failed.'})
        elif payment.status == 'cancelled':
            return JsonResponse({'status': 'cancelled'})

        result      = mpesa.stk_query(checkout_id)
        result_code = result.get('ResultCode')

        if result_code == '0':
            _activate_subscription(request.user, payment.plan, payment.billing_cycle, payment.amount)
            receipt        = result.get('MpesaReceiptNumber', '')
            payment.status = 'completed'
            payment.mpesa_receipt = receipt
            payment.save()
            return JsonResponse({'status': 'success', 'receipt': receipt})
        elif result_code == '1032':
            payment.status = 'cancelled'
            payment.save()
            return JsonResponse({'status': 'cancelled'})
        elif result_code is not None and result_code != '':
            payment.status         = 'failed'
            payment.failure_reason = result.get('ResultDesc', 'Payment failed.')
            payment.save()
            return JsonResponse({'status': 'failed', 'error': payment.failure_reason})

        return JsonResponse({'status': 'pending'})

    except Exception:
        return JsonResponse({'status': 'pending'})


def _activate_subscription(user, plan, billing_cycle, amount):
    from .models import UserSubscription
    from django.utils import timezone
    from datetime import timedelta

    days    = 365 if billing_cycle == 'annual' else 30
    expires = timezone.now() + timedelta(days=days)

    UserSubscription.objects.update_or_create(
        user=user,
        defaults={
            'plan': plan,
            'billing_cycle': billing_cycle,
            'amount_paid': amount,
            'expires_at': expires,
            'is_active': True,
        }
    )